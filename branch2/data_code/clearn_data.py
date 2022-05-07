#! /usr/bin/env python3
# -*- coding:utf-8 -*-
#  主要是为了将label excel表格和文件数据一一对齐
from __future__ import division, print_function
import numpy as np
import os
from openpyxl import load_workbook
from openpyxl import Workbook
import shutil

def read_xlxs(filename):
##first read the given labels information
    wb = load_workbook(filename)
    sheetnames = wb.get_sheet_names()
    ws = wb.get_sheet_by_name(sheetnames[0])
    # get the name of the first worksheets
    rows = ws.rows
    rows = list(rows)
    content = []
    for row in rows:
        line = [col.value for col in row]
        content.append(line)
    content = content[1:] 
    return content
     
filename = '/home/wangya/wine/distributed/data/files_1/data.xlsx'
file_dir = '/home/wangya/wine/distributed/data/files_1/data'
content1 = read_xlxs(filename)
# print(len(content1),len(content1[0]))
# print(content1[0][0])
xlsx_id = [i[0] for i in content1]
xlsx_id = set(xlsx_id)
file_id = [i.split('.')[0] for i in  os.listdir(file_dir)]
file_id = set(file_id)


xlsx_more_id = xlsx_id - file_id
print(xlsx_more_id)

file_more_id = file_id-xlsx_id
print(file_more_id)