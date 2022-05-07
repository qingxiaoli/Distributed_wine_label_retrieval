#! /usr/bin/env python3
# -*- coding:utf-8 -*-
from __future__ import division, print_function
import numpy as np
import os
from openpyxl import load_workbook
from openpyxl import Workbook
import shutil
import pickle

def read_xlxs(filename):
##first read the given labels information
    wb = load_workbook(filename)
    sheetnames = wb.get_sheet_names()
    ws = wb.get_sheet_by_name(sheetnames[0])
    # get the name of the first worksheets
    rows = ws.rows
    rows = list(rows)
    content = []
    for row in rows:
        line = [col.value for col in row]
        content.append(line)
    return content

def create_dictionary(content):
    label_dic = {}
    num_primary = 0
    for t in content[1:]:
        if t[1] in label_dic.keys():
            label_dic[t[1]].append(t)
        else:
            label_dic[t[1]] = []
            label_dic[t[1]].append(t)
            num_primary += 1

    num_data_primary = np.zeros(num_primary)
    name_primary = []
    order = 0
    for key in label_dic:
        name_primary.append(key)
        num_data_primary[order] = len(label_dic[key])
        order += 1
    sorted_order = np.argsort(num_data_primary)
    ordered_name_primary = []
    new_content = []
    for i in sorted_order:
        new_content.append([])
        ordered_name_primary.append(name_primary[i])
        for wine_info in label_dic[name_primary[i]]:
            new_content[-1].append(wine_info)
    num_data_primary = num_data_primary[sorted_order]

    return label_dic, ordered_name_primary, num_data_primary, new_content

def copy_data(label_dic, raw_files, subset, name_primary, subset_primary):
    for index in subset_primary:
        for file in label_dic[name_primary[index]]:
            filename = os.path.join(raw_files, str(file[0]) + '.jpg')
            if os.path.isfile(filename):
                shutil.copyfile(filename, os.path.join(subset, str(file[0]) + '.jpg'))

def write_xlsx(subset_content,xlsxfile_subset):
    wb=Workbook()
    sheet=wb.active 
    subset_content=[[['item_id','manufacturer','item_title','year']]]+subset_content
    row_num = 0
    for brand in range(len(subset_content)):  #
        for picture in range(len(subset_content[brand])):
            row_num += 1
            for item in range(len(subset_content[brand][picture])):
                sheet.cell(row=row_num, column=item + 1, value=str(subset_content[brand][picture][item]))
    wb.save(filename=xlsxfile_subset)

def create_dic(content):
    dic = {}
    for image in content[1:]:
        dic[image[0]] = [image[1], image[2]]
    return dic

if __name__ == '__main__':

    train_label_file = '/home/wangya/wine/distributed/data/files_2/train.xlsx'

    # create a dictionary containing the primary classes, the pictures belonging to it
    #  and the number of these pictures
    content = read_xlxs(train_label_file)
    dic = create_dic(content)
    with open('/home/wangya/wine/distributed/data/files_2/train_dic.pkl', 'wb') as point:
        pickle.dump(dic, point, protocol=2)
    print('dictionary is created!!')