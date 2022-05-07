#! /usr/bin/env python3
# -*- coding:utf-8 -*-
##下面是对于原来版本data下的subset去掉对应小于10样本的类并存为cut10_pics里面
from __future__ import division, print_function
import numpy as np
import os
from openpyxl import load_workbook
from openpyxl import Workbook
import shutil

def read_xlxs(filename):
##first read the given labels information
    wb = load_workbook(filename)
    sheetnames = wb.get_sheet_names()
    ws = wb.get_sheet_by_name(sheetnames[0])
    # get the name of the first worksheets
    rows = ws.rows
    rows = list(rows)
    content = []
    for row in rows:
        line = [col.value for col in row]
        content.append(line)
    return content

def create_dictionary(content):
    label_dic = {}
    num_primary = 0
    for t in content[1:]:
        if t[1] in label_dic.keys():
            label_dic[t[1]].append(t)
        else:
            label_dic[t[1]] = []
            label_dic[t[1]].append(t)
            num_primary += 1

    num_data_primary = np.zeros(num_primary)
    name_primary = []
    order = 0
    for key in label_dic:
        name_primary.append(key)
        num_data_primary[order] = len(label_dic[key])
        order += 1
    sorted_order = np.argsort(num_data_primary)
    ordered_name_primary = []
    new_content = []
    for i in sorted_order:
        new_content.append([])
        ordered_name_primary.append(name_primary[i])
        for wine_info in label_dic[name_primary[i]]:
            new_content[-1].append(wine_info)
    num_data_primary = num_data_primary[sorted_order]

    return label_dic, ordered_name_primary, num_data_primary, new_content

def copy_data(label_dic, raw_files, subset, name_primary, subset_primary):
    for index in subset_primary:
        for file in label_dic[name_primary[index]]:
            filename = os.path.join(raw_files, str(file[0]) + '.jpg')
            if os.path.isfile(filename):
                shutil.copyfile(filename, os.path.join(subset, str(file[0]) + '.jpg'))

def write_xlsx(subset_content,xlsxfile_subset):
    wb=Workbook()
    sheet=wb.active 
    subset_content=[[['item_id','manufacturer','item_title','year']]]+subset_content
    row_num = 0
    for brand in range(len(subset_content)):  #
        for picture in range(len(subset_content[brand])):
            row_num += 1
            for item in range(len(subset_content[brand][picture])):
                sheet.cell(row=row_num, column=item + 1, value=str(subset_content[brand][picture][item]))
    wb.save(filename=xlsxfile_subset)

if __name__ == '__main__':

    r80_label_file = '../data/bzxx80.xlsx'
    r80w_files = '../data/80w_pics/'
    cut10_label_file = '../data/cut10.xlsx'
    cut10_files = '../data/cut10_pics/'

    # create a dictionary containing the primary classes, the pictures belonging to it
    #  and the number of these pictures
    content = read_xlxs(cut10_label_file)
    print('content is created!!')
    label_dic, name_primary, num_data_primary, sorted_content = create_dictionary(content)
    print('sorted_content is created!!')

    # choose the data sets which contain more than 10 samples
    total_primary = range(len(num_data_primary))
    break_point1 = np.where(num_data_primary <= 10)
    break_point1 = break_point1[0][-1]
    cut10_num_data_primary = num_data_primary[break_point1 + 1 :]
    cut10_name_primary = name_primary[break_point1 + 1 :]
    cut10_primary = total_primary[break_point1 + 1 :]
    cut10_sorted_content = sorted_content[break_point1 + 1 :]

    copy_data(label_dic, r80w_files, cut10_files, name_primary, cut10_primary)
    print('lists corresponding to all subsets is created!!')
    write_xlsx(cut10_sorted_content, cut10_label_file)
    print('congratulation! {}is written!'.format(cut10_label_file))