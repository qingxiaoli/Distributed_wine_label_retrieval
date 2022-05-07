from PIL import Image
import random
from openpyxl import load_workbook
from openpyxl import Workbook
import os
import shutil

def read_xlxs(filename):
##first read the given labels information 
    wb = load_workbook(filename)
    sheetnames = wb.get_sheet_names()  
    ws = wb.get_sheet_by_name(sheetnames[0]) 
    # get the name of the first worksheets
    rows = ws.rows
    rows = list(rows)
    content = []
    for row in rows:
        line = [col.value for col in row]
        content.append(line)
    return content


def compute_big_class(filename):
    content = read_xlxs(filename)
    class_list = [t[1] for t in content[1:]] #shengchanchangjia is the second col
    
    #first delete repetition in class_list
    class_list = list(set(class_list))
    big_class_num = len(class_list)
    return big_class_num
   
def compute_small_class(filename):
    content = read_xlxs(filename)
    class_list = [t[2] for t in content[1:]] #item_text is the third col
    
    #first delete repetition in class_list
    class_list = list(set(class_list))
    small_class_num = len(class_list)
    return small_class_num
    
 
##the follows are compute all class in all 80wdata 
filename =  r'/home/wangya/wine/distributed_4/data/files_4/train.xlsx'
big_class_num  =  compute_big_class(filename)
small_class_num  =  compute_small_class(filename)
print('the big_class_num of all data is:{0}'.format(big_class_num))
print('the small_class_num of all data is:{0}'.format(small_class_num))


