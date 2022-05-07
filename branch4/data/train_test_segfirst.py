#! /usr/bin/env python3
# -*- coding:utf-8 -*-
from __future__ import division, print_function
import numpy as np
import os
from openpyxl import load_workbook
from openpyxl import Workbook
import shutil

def read_xlxs(filename):
##first read the given labels information
    wb = load_workbook(filename)
    sheetnames = wb.get_sheet_names()
    ws = wb.get_sheet_by_name(sheetnames[0])
    # get the name of the first worksheets
    rows = ws.rows
    rows = list(rows)
    content = []
    for row in rows:
        line = [col.value for col in row]
        content.append(line)
    return content

def create_dictionary(content):
    label_dic = {}
    num_primary = 0
    for t in content[1:]:
        if t[1] in label_dic.keys():
            label_dic[t[1]].append(t)
        else:
            label_dic[t[1]] = []
            label_dic[t[1]].append(t)
            num_primary += 1

    num_data_primary = np.zeros(num_primary)
    name_primary = []
    order = 0
    for key in label_dic:
        name_primary.append(key)
        num_data_primary[order] = len(label_dic[key])
        order += 1
    sorted_order = np.argsort(num_data_primary)
    ordered_name_primary = []
    new_content = []
    for i in sorted_order:
        new_content.append([])
        ordered_name_primary.append(name_primary[i])
        for wine_info in label_dic[name_primary[i]]:
            new_content[-1].append(wine_info)
    num_data_primary = num_data_primary[sorted_order]

    return label_dic, ordered_name_primary, num_data_primary, new_content

def seg_train_test(num_data_primary, sorted_content, test_num):
    train_cont = [[] for _ in range(len(sorted_content))]
    test_cont = [[] for _ in range(len(sorted_content))]
    for i in range(len(test_num)):
        assert test_num[i] <= num_data_primary[i]
        candidate = list(range(num_data_primary[i]))
        candidate2 = list(range(num_data_primary[i]))
        subbrand_num_dic = {}
        for k in range(len(sorted_content[i])):
            if sorted_content[i][k][2] in subbrand_num_dic.keys():
                subbrand_num_dic[sorted_content[i][k][2]] += 1
            else:
                subbrand_num_dic[sorted_content[i][k][2]] = 1

        for j in range(test_num[i]):
            while candidate:
                chosen = np.random.choice(candidate, 1)[0]
                if subbrand_num_dic[sorted_content[i][chosen][2]] >= 2:
                    test_cont[i].append(sorted_content[i][chosen])
                    candidate.remove(chosen)
                    candidate2.remove(chosen)
                    subbrand_num_dic[sorted_content[i][chosen][2]] -= 1
                    break
                else:
                    candidate.remove(chosen)
        del subbrand_num_dic
        for j in candidate2:
            train_cont[i].append(sorted_content[i][j])
        pass
    pass
    return train_cont, test_cont

def copy_data_from_content(raw_files, save_file, content):
    for brand in content:
        for picture in brand:
            filename = os.path.join(raw_files, str(picture[0]) + '.jpg')
            if os.path.isfile(filename):
                shutil.copyfile(filename, os.path.join(save_file, str(picture[0]) + '.jpg'))

def write_xlsx(subset_content, xlsxfile_subset):
    wb=Workbook()
    sheet=wb.active 
    subset_content=[[['item_id','manufacturer','item_title','year']]]+subset_content
    row_num = 0
    for brand in range(len(subset_content)):  #
        for picture in range(len(subset_content[brand])):
            row_num += 1
            for item in range(len(subset_content[brand][picture])):
                sheet.cell(row=row_num, column=item + 1, value=str(subset_content[brand][picture][item].encode('utf-8')))
    wb.save(filename=xlsxfile_subset)

if __name__ == '__main__':

    cut10_label_file = '/home/wangya/wine/distributed_4/data/files_4/data.xlsx'
    cut10_files = '/home/wangya/wine/distributed_4/data/files_4/data/'
    train_label_file = '/home/wangya/wine/distributed_4/data/files_4/train.xlsx'
    train_files = '/home/wangya/wine/distributed_4/data/files_4/train/'
    test_label_file = '/home/wangya/wine/distributed_4/data/files_4/test.xlsx'
    test_files = '/home/wangya/wine/distributed_4/data/files_4/test/'

    # create a dictionary containing the primary classes, the pictures belonging to it
    #  and the number of these pictures
    content = read_xlxs(cut10_label_file)
    print('content is created!!')
    label_dic, name_primary, num_data_primary, sorted_content = create_dictionary(content)
    print('sorted_content is created!!')

    # determine the numbef of the training pictures and the testing pictures
    pro = 0.20
    test_num = list(map(lambda t: int(t * pro), num_data_primary))
    num_data_primary = list(map(lambda t: int(t), num_data_primary))

    # create the infromation of training pictures and the testing pictures
    train_cont, test_cont = seg_train_test(num_data_primary, sorted_content, test_num)

    # copy file from 'cut10_pics' to 'train' and 'test'
    copy_data_from_content(cut10_files, train_files, train_cont)
    print('training file is copyed!')
    copy_data_from_content(cut10_files, test_files, test_cont)
    print('testing file is copyed!')

    # create xlsx files which contain labels information
    write_xlsx(train_cont, train_label_file)
    print('training xlsx file is created!')
    write_xlsx(test_cont, test_label_file)
    print('testing xlsx file is created!')
