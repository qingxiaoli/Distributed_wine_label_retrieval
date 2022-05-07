# -*- coding: utf-8 -*-
##本code主要是自动对算法生成的结果进行评估打分

from __future__ import division, print_function
import numpy as np
import os
from openpyxl import load_workbook
from openpyxl import Workbook
import shutil
# import pdb

def read_xlxs(filename):
##first read the given labels information
    wb = load_workbook(filename)
    sheetnames = wb.get_sheet_names()
    ws = wb.get_sheet_by_name(sheetnames[0])
    # get the name of the first worksheets
    rows = ws.rows
    rows = list(rows)
    content = []
    for row in rows:
        line = [col.value for col in row]
        content.append(line)
    return content

def creat_dic(content):
    cont_dic = {}
    cont = content
    del cont[0]
    for item in cont:
        cont_dic[str(item[0])] = item[1:3]
#print(cont_dic[item[0]])
    return cont_dic

def compare(test_id,result_id,test_dic,train_dic):
    b=0
    i=0
# test_id = test_img.split('.')[0]
# result_id = result_img.split('.')[0]
    # print(test_dic.get(test_id))
    # print(test_dic[test_id])
    test_brand = test_dic[test_id][0]
    result_brand = train_dic[result_id][0]
    test_item = test_dic[test_id][1]
    result_item = train_dic[result_id][1]
    if test_brand == result_brand:
        b = b+1
        notright1 = 0
        notright2 = 0
        test_item_list = test_item.split()
        result_item_list = result_item.split()
        for test_il in test_item_list:
            if test_il not in result_item_list:
                notright1 = 1
                break
        for result_il in result_item_list:
            if result_il not in test_item_list:
                notright2 = 1
                break
        if notright1 == 0 or notright2 == 0:
            i = i+1
    # if i==0:
    #     # print(test_id)
    #     # print(test_item)
        # print(result_item)
        # print('\n')
        # os.mkdir(str(test_id))
        # shutil.copy()
    return b,i

train_xlxs = '/home/wangya/wine/distributed_4/data/files_4/train.xlsx'
test_xlxs = '/home/wangya/wine/distributed_4/data/files_4/test.xlsx'
train_cont = read_xlxs(train_xlxs)

test_cont = read_xlxs(test_xlxs)
train_dic = creat_dic(train_cont)
test_dic = creat_dic(test_cont)
tests_id = [t[0] for t in test_cont] 
results_dir = './result_5brand_5'
#for file in os.listdir(results_dir):
bb=0
ii=0
num = len(os.listdir(results_dir))
print(num)
for di in os.listdir(results_dir):

    result_dir = os.path.join(results_dir,di)
    #print(result_dir)
    files = os.listdir(result_dir)
    
    for i in range(len(files)):
        file_id = files[i].split('.')[0]
        if file_id in tests_id:
            test_id = file_id
            results_id = []
            for j in range(len(files)):
                if i != j:
                    results_id.append(files[j].split('.')[0])
        else:
            continue

    # print(test_id)
    # print(result_id)
    i = 0
    b = 0
    for result_id in results_id:
        bt,it = compare(test_id,result_id,test_dic,train_dic)
        if bt != 0:
            b = 1
        if it != 0:
            i = 1
    bb += b
    ii += i
print('brand_acc is ',bb/num)
print('item_acc is',ii/num)

